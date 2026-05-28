import os
import re
import openpyxl
import subprocess
import sys

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
POPULAR_XLSX = os.path.join(BASE_DIR, "Cancionero Popular", "Cancionero_popular_Lista_canciones.xlsx")
DIOS_XLSX = os.path.join(BASE_DIR, "Canciones para Dios Word", "Canciones_Dios_Lista_canciones.xlsx")
ACORDES_HTML = os.path.join(BASE_DIR, "acordes.html")
RAW_TXT = os.path.join(BASE_DIR, "cancionero_raw.txt")
PDF_TO_HTML_PY = os.path.join(BASE_DIR, "pdf_to_html.py")

def parse_excel_songs(filepath, default_col_mapping):
    if not os.path.exists(filepath):
        print(f"Advertencia: Archivo de Excel no encontrado en {filepath}")
        return []
        
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active
    
    header_row_idx = None
    cols_map = {}
    
    # Escanear las primeras 10 filas para encontrar cabeceras
    for r in range(1, min(11, sheet.max_row + 1)):
        row_vals = [str(sheet.cell(row=r, column=c).value or '').strip().lower() for c in range(1, sheet.max_column + 1)]
        if any(keyword in row_vals for keyword in ["titulo", "numero", "nº", "n°", "tutorial", "cover", "covers"]):
            header_row_idx = r
            break
            
    if header_row_idx:
        # Primera pasada: Coincidencia exacta de cabeceras
        for c in range(1, sheet.max_column + 1):
            val = str(sheet.cell(row=header_row_idx, column=c).value or '').strip().lower()
            if val in ["numero", "número", "num", "nº", "n°", "id"]:
                cols_map['id'] = c
            elif val in ["titulo", "título", "cancion", "canción", "tema"]:
                cols_map['title'] = c
            elif val in ["tutorial", "tutoriales"]:
                cols_map['tutorial'] = c
            elif val in ["covers", "cover"]:
                cols_map['cover'] = c
                
        # Segunda pasada: Búsqueda flexible por subcadena
        for c in range(1, sheet.max_column + 1):
            val = str(sheet.cell(row=header_row_idx, column=c).value or '').strip().lower()
            if 'id' not in cols_map and any(k in val for k in ["numero", "num", "nº", "n°", "id"]):
                cols_map['id'] = c
            if 'title' not in cols_map and any(k in val for k in ["titulo", "cancion", "tema"]):
                cols_map['title'] = c
            if 'tutorial' not in cols_map and "tutorial" in val and "list" not in val:
                cols_map['tutorial'] = c
            if 'cover' not in cols_map and "cover" in val and "list" not in val:
                cols_map['cover'] = c
                
        start_row = header_row_idx + 1
    else:
        # Sin cabeceras, usar mapeo predeterminado
        cols_map = default_col_mapping.copy()
        
        # Buscar columnas de YouTube dinámicamente
        for c in range(1, sheet.max_column + 1):
            for r in range(1, min(20, sheet.max_row + 1)):
                val = str(sheet.cell(row=r, column=c).value or '')
                if "youtu" in val and "results?" not in val:
                    if 'tutorial' not in cols_map:
                        cols_map['tutorial'] = c
                    elif cols_map['tutorial'] != c and 'cover' not in cols_map:
                        cols_map['cover'] = c
                    break
        start_row = 1
        
    songs = []
    for r in range(start_row, sheet.max_row + 1):
        id_val = sheet.cell(row=r, column=cols_map.get('id', 1)).value
        title_val = sheet.cell(row=r, column=cols_map.get('title', 2)).value
        
        if id_val is None and title_val is None:
            continue
            
        tutorial_val = sheet.cell(row=r, column=cols_map.get('tutorial', 99)).value if 'tutorial' in cols_map else None
        cover_val = sheet.cell(row=r, column=cols_map.get('cover', 99)).value if 'cover' in cols_map else None
        
        title_str = str(title_val).strip() if title_val is not None else ''
        id_str = str(int(id_val)) if isinstance(id_val, (int, float)) else (str(id_val).strip() if id_val is not None else '')
        if id_str.endswith('.0'):
            id_str = id_str[:-2]
            
        # Omitir si la fila es otra cabecera
        if title_str.lower() in ['titulo', 'título del canto', 'título', 'tema', 'canción', 'none', 'nan', ''] or id_str.lower() in ['numero', 'número', 'nº', 'n°', 'id', 'none', 'nan', '']:
            continue
            
        songs.append({
            'id': id_str,
            'title': title_str,
            'tutorial': str(tutorial_val).strip() if tutorial_val is not None else '',
            'cover': str(cover_val).strip() if cover_val is not None else ''
        })
        
    return songs

def extract_current_cancionero(acordes_path):
    if not os.path.exists(acordes_path):
        print(f"Error: {acordes_path} no encontrado.")
        return {"populares": {}, "dios": {}}
        
    with open(acordes_path, 'r', encoding='utf-8') as f:
        content = f.read()
        
    match = re.search(r'const cancionero = (\{.*?\});', content, re.DOTALL)
    if not match:
        print("No se encontró const cancionero en acordes.html")
        return {"populares": {}, "dios": {}}
        
    db_str = match.group(1)
    
    # Encontrar posiciones de bloques populares y dios
    pop_search = re.search(r'populares:\s*\{', db_str)
    dios_search = re.search(r'dios:\s*\{', db_str)
    
    if not pop_search or not dios_search:
        print("No se encontraron las categorías populares/dios dentro de cancionero.")
        return {"populares": {}, "dios": {}}
        
    pop_start = pop_search.end()
    dios_start = dios_search.end()
    
    pop_content = db_str[pop_start : db_str.find('dios:', pop_start)]
    dios_content = db_str[dios_start : db_str.rfind('}')]
    
    cancionero = {"populares": {}, "dios": {}}
    song_pattern = r'"([^"]+)":\s*\{(.*?)\}'
    
    # Parsear populares
    for m in re.finditer(song_pattern, pop_content):
        name = m.group(1)
        fields_str = m.group(2)
        def get_field(field_name):
            pattern = fr'{field_name}:\s*(["\'])(.*?)\1'
            field_match = re.search(pattern, fields_str)
            return field_match.group(2).strip() if field_match else ""
        cancionero["populares"][name] = {
            "pdf": get_field("pdf"),
            "tutorial": get_field("tutorial"),
            "cover": get_field("cover"),
            "description": get_field("description")
        }
        
    # Parsear dios
    for m in re.finditer(song_pattern, dios_content):
        name = m.group(1)
        fields_str = m.group(2)
        def get_field(field_name):
            pattern = fr'{field_name}:\s*(["\'])(.*?)\1'
            field_match = re.search(pattern, fields_str)
            return field_match.group(2).strip() if field_match else ""
        cancionero["dios"][name] = {
            "pdf": get_field("pdf"),
            "tutorial": get_field("tutorial"),
            "cover": get_field("cover"),
            "description": get_field("description")
        }
        
    return cancionero

def format_cancionero_js_string(cancionero):
    lines = []
    lines.append("{")
    
    for cat in ["populares", "dios"]:
        lines.append(f"            {cat}: {{")
        cat_data = cancionero[cat]
        sorted_keys = sorted(cat_data.keys(), key=lambda s: s.lower())
        
        for i, key in enumerate(sorted_keys):
            song = cat_data[key]
            pdf = song.get('pdf', '')
            tutorial = song.get('tutorial', '')
            cover = song.get('cover', '')
            description = song.get('description', '')
            
            desc_escaped = description.replace('\\', '\\\\').replace('"', '\\"').replace('\n', '\\n')
            
            line = f'                "{key}": {{ pdf: "{pdf}", tutorial: "{tutorial}", cover: "{cover}"'
            if desc_escaped:
                line += f', description: "{desc_escaped}"'
            line += ' }'
            
            if i < len(sorted_keys) - 1:
                line += ','
            lines.append(line)
            
        if cat == "populares":
            lines.append("            },")
        else:
            lines.append("            }")
            
    lines.append("        }")
    return "\n".join(lines)

def main():
    print("Iniciando sincronización de canciones desde Excel...")
    
    # 1. Parsear canciones del Excel
    print("Leyendo archivos Excel...")
    pop_excel = parse_excel_songs(POPULAR_XLSX, {'id': 1, 'title': 2, 'tutorial': 3, 'cover': 6})
    dios_excel = parse_excel_songs(DIOS_XLSX, {'id': 1, 'title': 2})
    
    print(f"Leídas {len(pop_excel)} canciones populares del Excel.")
    print(f"Leídas {len(dios_excel)} canciones religiosas del Excel.")
    
    # 2. Cargar base de datos actual de acordes.html
    print("Cargando cancionero actual de acordes.html...")
    existing_db = extract_current_cancionero(ACORDES_HTML)
    
    print(f"Base de datos actual: {len(existing_db['populares'])} populares, {len(existing_db['dios'])} religiosas.")
    
    # 3. Combinar datos
    new_db = {"populares": {}, "dios": {}}
    
    # Combinar populares
    for song in pop_excel:
        title_raw = song['title']
        clean_title = title_raw.split('(')[0].strip()
        pdf_name = f"{title_raw}{song['id']}.pdf"
        
        description = ""
        # Preservar descripción si ya existía
        if clean_title in existing_db['populares']:
            description = existing_db['populares'][clean_title].get('description', '')
            
        new_db['populares'][clean_title] = {
            'pdf': pdf_name,
            'tutorial': song['tutorial'],
            'cover': song['cover'],
            'description': description
        }
        
    # Combinar dios
    for song in dios_excel:
        title_raw = song['title']
        clean_title = title_raw.strip()
        pdf_name = f"{title_raw}{song['id']}.pdf"
        
        description = ""
        if clean_title in existing_db['dios']:
            description = existing_db['dios'][clean_title].get('description', '')
            
        new_db['dios'][clean_title] = {
            'pdf': pdf_name,
            'tutorial': song['tutorial'],
            'cover': song['cover'],
            'description': description
        }
        
    print(f"Base de datos unificada: {len(new_db['populares'])} populares, {len(new_db['dios'])} religiosas.")
    
    # 4. Generar el bloque JS cancionero
    raw_js = format_cancionero_js_string(new_db)
    
    # 5. Actualizar acordes.html
    print("Actualizando acordes.html...")
    with open(ACORDES_HTML, 'r', encoding='utf-8') as f:
        html_content = f.read()
        
    start_idx = html_content.find('const cancionero = {')
    if start_idx == -1:
        start_idx = html_content.find('const cancionero={')
        
    if start_idx != -1:
        end_marker = "let currentCategory = 'populares';"
        end_idx = html_content.find(end_marker, start_idx)
        if end_idx != -1:
            block_end = html_content.rfind('};', start_idx, end_idx) + 2
            new_declaration = f"const cancionero = {raw_js};\n\n        "
            updated_html = html_content[:start_idx] + new_declaration + html_content[block_end:]
            
            with open(ACORDES_HTML, 'w', encoding='utf-8') as f:
                f.write(updated_html)
            print("acordes.html actualizado con éxito.")
        else:
            print("Error: No se pudo encontrar el marcador de fin del bloque cancionero en acordes.html.")
            sys.exit(1)
    else:
        print("Error: No se pudo encontrar 'const cancionero = {' en acordes.html.")
        sys.exit(1)
        
    # 6. Actualizar cancionero_raw.txt
    print("Actualizando cancionero_raw.txt...")
    with open(RAW_TXT, 'w', encoding='utf-8') as f:
        f.write(raw_js)
    print("cancionero_raw.txt actualizado con éxito.")
    
    # 7. Ejecutar pdf_to_html.py para recompilar el visor
    if os.path.exists(PDF_TO_HTML_PY):
        print("Ejecutando pdf_to_html.py para actualizar letras, sitemap y songs_data.js...")
        try:
            result = subprocess.run([sys.executable, PDF_TO_HTML_PY], capture_output=True, text=True, encoding='latin-1', errors='replace')
            print("--- SALIDA DE pdf_to_html.py ---")
            print(result.stdout)
            if result.stderr:
                print("--- ERRORES DE pdf_to_html.py ---")
                print(result.stderr)
        except Exception as e:
            print(f"Error al ejecutar pdf_to_html.py: {e}")
    else:
        print("Advertencia: pdf_to_html.py no encontrado.")

if __name__ == "__main__":
    main()
